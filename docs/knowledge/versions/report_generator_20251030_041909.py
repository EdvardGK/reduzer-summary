# -*- coding: utf-8 -*-
"""
Enhanced Report Generation for LCA Scenario Mapping
Generates Excel and PDF reports with visualizations and insights
"""

import io
import pandas as pd
from datetime import datetime
from typing import Dict, Any, Optional, List, Tuple

# PDF imports
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, Image
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY
from reportlab.pdfgen import canvas
import plotly.graph_objects as go

# Excel charting
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


# ==============================================================================
# INSIGHT GENERATION
# ==============================================================================

def generate_insights(df: pd.DataFrame, structure: Dict[str, Any],
                     scenario_summary: pd.DataFrame) -> Dict[str, Any]:
    """
    Generate key insights from the data.

    Returns:
        Dictionary with insights, findings, and recommendations
    """
    insights = {
        'executive_summary': [],
        'key_findings': [],
        'recommendations': []
    }

    # Mapping completeness
    total_rows = len(df)
    mapped_rows = len(df[~df['excluded'] & df['mapped_scenario'].notna()])
    excluded_rows = int(df['excluded'].sum())
    completeness = (mapped_rows / total_rows * 100) if total_rows > 0 else 0

    insights['executive_summary'].append(
        f"Analysen dekker {total_rows} datapunkter, hvor {mapped_rows} ({completeness:.0f}%) "
        f"er kartlagt og {excluded_rows} er ekskludert."
    )

    # Scenario comparison
    if 'C' in structure and 'A' in structure:
        from .data_parser import compare_scenarios
        comparison = compare_scenarios(structure, 'A', 'C')

        if comparison:
            ratio = comparison['ratio']['total_gwp']
            diff = comparison['difference']['total_gwp']

            if ratio < 90:
                verdict = "betydelig reduksjon"
                color = "green"
            elif ratio < 100:
                verdict = "moderat reduksjon"
                color = "green"
            elif ratio < 110:
                verdict = "liten økning"
                color = "yellow"
            else:
                verdict = "betydelig økning"
                color = "red"

            insights['executive_summary'].append(
                f"Scenario C viser en {verdict} i klimagassutslipp sammenlignet med Scenario A "
                f"({ratio:.1f}%, {diff:+,.0f} kg CO2e)."
            )

            insights['key_findings'].append({
                'title': 'Scenario C vs A Sammenligning',
                'value': f"{ratio:.1f}%",
                'description': f"Differanse: {diff:+,.0f} kg CO2e",
                'color': color
            })

            # Phase analysis
            for key, label in [
                ('construction_a', 'Konstruksjon (A)'),
                ('operation_b', 'Drift (B)'),
                ('end_of_life_c', 'Avslutning (C)')
            ]:
                phase_ratio = comparison['ratio'][key]
                phase_diff = comparison['difference'][key]

                if phase_ratio and abs(100 - phase_ratio) > 10:
                    change = "reduksjon" if phase_ratio < 100 else "økning"
                    insights['key_findings'].append({
                        'title': f'{label}: {change.capitalize()}',
                        'value': f"{phase_ratio:.1f}%",
                        'description': f"{phase_diff:+,.0f} kg CO2e",
                        'color': 'green' if phase_ratio < 100 else 'red'
                    })

    # MMI distribution insights
    for scenario in structure.keys():
        from .visualizations import get_mmi_summary_stats
        mmi_stats = get_mmi_summary_stats(structure, scenario)

        if not mmi_stats.empty:
            # Find dominant MMI category
            dominant = mmi_stats.iloc[0]
            dominant_pct = dominant['Andel (%)']

            insights['key_findings'].append({
                'title': f'Scenario {scenario}: Dominerende MMI',
                'value': f"{dominant['MMI']} - {dominant['Status']}",
                'description': f"{dominant_pct:.1f}% av total GWP ({dominant['GWP (kg CO2e)']:,.0f} kg CO2e)",
                'color': 'blue'
            })

    # Discipline analysis
    if structure:
        for scenario, scenario_data in structure.items():
            disciplines = scenario_data['disciplines']
            if len(disciplines) > 0:
                # Find largest discipline contributor
                disc_totals = {disc: data['total']['total_gwp']
                              for disc, data in disciplines.items()}
                largest_disc = max(disc_totals, key=disc_totals.get)
                largest_gwp = disc_totals[largest_disc]
                scenario_total = scenario_data['total']['total_gwp']
                pct = (largest_gwp / scenario_total * 100) if scenario_total > 0 else 0

                insights['key_findings'].append({
                    'title': f'Scenario {scenario}: Største bidragsyter',
                    'value': largest_disc,
                    'description': f"{pct:.1f}% av total ({largest_gwp:,.0f} kg CO2e)",
                    'color': 'blue'
                })

    # Recommendations
    if 'C' in structure and 'A' in structure:
        comparison = compare_scenarios(structure, 'A', 'C')
        if comparison and comparison['ratio']['total_gwp'] > 100:
            insights['recommendations'].append(
                "Scenario C har høyere klimagassutslipp enn Scenario A. Vurder alternative løsninger "
                "for å redusere utslipp, spesielt i faser med størst økning."
            )
        elif comparison and comparison['ratio']['total_gwp'] < 90:
            insights['recommendations'].append(
                "Scenario C viser god reduksjon i klimagassutslipp. Dokumenter og implementer "
                "disse tiltakene for maksimal klimanytte."
            )

    if completeness < 90:
        insights['recommendations'].append(
            f"Kartleggingsgrad er {completeness:.0f}%. Vurder å kartlegge flere datapunkter "
            "for mer komplett analyse."
        )

    return insights


# ==============================================================================
# EXCEL EXPORT WITH VISUALIZATIONS
# ==============================================================================

def generate_excel_report(df: pd.DataFrame, structure: Dict[str, Any],
                         scenario_summary: pd.DataFrame) -> io.BytesIO:
    """
    Generate comprehensive Excel report with charts and insights.
    """
    excel_output = io.BytesIO()

    # Generate insights
    insights = generate_insights(df, structure, scenario_summary)

    with pd.ExcelWriter(excel_output, engine='openpyxl') as writer:
        workbook = writer.book

        # ==============================================================================
        # SHEET 1: EXECUTIVE SUMMARY
        # ==============================================================================
        ws_summary = workbook.create_sheet("📊 Sammendrag", 0)

        # Title
        ws_summary['A1'] = 'LCA SCENARIO MAPPING - SAMMENDRAG'
        ws_summary['A1'].font = Font(size=18, bold=True, color='1565C0')
        ws_summary.merge_cells('A1:D1')
        ws_summary['A1'].alignment = Alignment(horizontal='center')

        ws_summary['A2'] = f"Generert: {datetime.now().strftime('%d.%m.%Y kl. %H:%M')}"
        ws_summary['A2'].alignment = Alignment(horizontal='center')
        ws_summary.merge_cells('A2:D2')

        row = 4

        # Executive Summary
        ws_summary[f'A{row}'] = 'HOVEDFUNN'
        ws_summary[f'A{row}'].font = Font(size=14, bold=True, color='1565C0')
        row += 1

        for summary in insights['executive_summary']:
            ws_summary[f'A{row}'] = summary
            ws_summary[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
            ws_summary.row_dimensions[row].height = 30
            ws_summary.merge_cells(f'A{row}:D{row}')
            row += 1

        row += 1

        # Key Findings
        ws_summary[f'A{row}'] = 'NØKKELTALL'
        ws_summary[f'A{row}'].font = Font(size=14, bold=True, color='1565C0')
        row += 1

        ws_summary[f'A{row}'] = 'Kategori'
        ws_summary[f'B{row}'] = 'Verdi'
        ws_summary[f'C{row}'] = 'Detaljer'
        for col in ['A', 'B', 'C']:
            ws_summary[f'{col}{row}'].font = Font(bold=True)
            ws_summary[f'{col}{row}'].fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
        row += 1

        for finding in insights['key_findings']:
            ws_summary[f'A{row}'] = finding['title']
            ws_summary[f'B{row}'] = finding['value']
            ws_summary[f'C{row}'] = finding['description']

            # Color coding
            if finding['color'] == 'green':
                fill_color = 'C8E6C9'
            elif finding['color'] == 'red':
                fill_color = 'FFCDD2'
            elif finding['color'] == 'yellow':
                fill_color = 'FFF9C4'
            else:
                fill_color = 'E3F2FD'

            ws_summary[f'B{row}'].fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
            ws_summary[f'B{row}'].font = Font(bold=True)
            row += 1

        row += 1

        # Recommendations
        if insights['recommendations']:
            ws_summary[f'A{row}'] = 'ANBEFALINGER'
            ws_summary[f'A{row}'].font = Font(size=14, bold=True, color='1565C0')
            row += 1

            for i, rec in enumerate(insights['recommendations'], 1):
                ws_summary[f'A{row}'] = f"{i}. {rec}"
                ws_summary[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
                ws_summary.row_dimensions[row].height = 40
                ws_summary.merge_cells(f'A{row}:D{row}')
                row += 1

        # Column widths
        ws_summary.column_dimensions['A'].width = 30
        ws_summary.column_dimensions['B'].width = 20
        ws_summary.column_dimensions['C'].width = 40
        ws_summary.column_dimensions['D'].width = 15

        # ==============================================================================
        # SHEET 2: SCENARIO SUMMARY WITH CHART
        # ==============================================================================
        scenario_summary.to_excel(writer, sheet_name='📈 Scenarioer', startrow=1, index=False)
        ws_scenarios = writer.sheets['📈 Scenarioer']

        # Title
        ws_scenarios['A1'] = 'SCENARIO-OPPSUMMERING'
        ws_scenarios['A1'].font = Font(size=14, bold=True, color='1565C0')

        # Header formatting
        for col in range(1, len(scenario_summary.columns) + 1):
            cell = ws_scenarios.cell(2, col)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='1565C0', end_color='1565C0', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')

        # Add bar chart
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "GWP per Scenario"
        chart.y_axis.title = 'kg CO2e'
        chart.x_axis.title = 'Scenario'

        data = Reference(ws_scenarios, min_col=2, min_row=2, max_row=len(scenario_summary)+2, max_col=5)
        cats = Reference(ws_scenarios, min_col=1, min_row=3, max_row=len(scenario_summary)+2)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 12
        chart.width = 20

        ws_scenarios.add_chart(chart, f"A{len(scenario_summary)+5}")

        # ==============================================================================
        # SHEET 3: C vs A COMPARISON
        # ==============================================================================
        if 'C' in structure and 'A' in structure:
            from .data_parser import compare_scenarios
            comparison = compare_scenarios(structure, 'A', 'C')

            if comparison:
                comp_data = []
                for key, label in [
                    ('total_gwp', 'Total GWP'),
                    ('construction_a', 'Konstruksjon (A)'),
                    ('operation_b', 'Drift (B)'),
                    ('end_of_life_c', 'Avslutning (C)')
                ]:
                    base_val = structure['A']['total'][key]
                    comp_val = structure['C']['total'][key]
                    diff = comparison['difference'][key]
                    ratio = comparison['ratio'][key] if comparison['ratio'][key] else 0

                    # Verdict
                    if ratio < 90:
                        verdict = "Stor reduksjon ✓"
                    elif ratio < 100:
                        verdict = "Reduksjon ✓"
                    elif ratio < 110:
                        verdict = "Liten økning ⚠"
                    else:
                        verdict = "Stor økning ✗"

                    comp_data.append({
                        'LCA-fase': label,
                        'Scenario A': base_val,
                        'Scenario C': comp_val,
                        'Differanse (C-A)': diff,
                        'Ratio (C/A) %': ratio,
                        'Vurdering': verdict
                    })

                comp_df = pd.DataFrame(comp_data)
                comp_df.to_excel(writer, sheet_name='🔄 C vs A', startrow=1, index=False)
                ws_comp = writer.sheets['🔄 C vs A']

                # Title
                ws_comp['A1'] = 'SCENARIO C vs A SAMMENLIGNING'
                ws_comp['A1'].font = Font(size=14, bold=True, color='1565C0')

                # Header formatting
                for col in range(1, len(comp_df.columns) + 1):
                    cell = ws_comp.cell(2, col)
                    cell.font = Font(bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color='1565C0', end_color='1565C0', fill_type='solid')

                # Conditional formatting on ratio column
                for row in range(3, len(comp_data) + 3):
                    ratio_cell = ws_comp.cell(row, 5)
                    ratio_val = ratio_cell.value

                    if ratio_val < 100:
                        ratio_cell.fill = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')
                    else:
                        ratio_cell.fill = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')

        # ==============================================================================
        # SHEET 4: MMI DISTRIBUTION
        # ==============================================================================
        from .visualizations import get_mmi_summary_stats
        mmi_dist_data = []
        for scenario in structure.keys():
            mmi_stats = get_mmi_summary_stats(structure, scenario)
            if not mmi_stats.empty:
                mmi_stats['Scenario'] = scenario
                mmi_dist_data.append(mmi_stats)

        if mmi_dist_data:
            mmi_dist_df = pd.concat(mmi_dist_data, ignore_index=True)
            mmi_dist_df = mmi_dist_df[['Scenario', 'MMI', 'Status', 'GWP (kg CO2e)', 'Andel (%)', 'Antall rader']]
            mmi_dist_df.to_excel(writer, sheet_name='🏗️ MMI Fordeling', startrow=1, index=False)

            ws_mmi = writer.sheets['🏗️ MMI Fordeling']
            ws_mmi['A1'] = 'MMI-FORDELING PER SCENARIO'
            ws_mmi['A1'].font = Font(size=14, bold=True, color='1565C0')

            # Header formatting
            for col in range(1, len(mmi_dist_df.columns) + 1):
                cell = ws_mmi.cell(2, col)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='1565C0', end_color='1565C0', fill_type='solid')

        # ==============================================================================
        # SHEET 5: DISCIPLINE BREAKDOWN
        # ==============================================================================
        discipline_data = []
        for scenario, scenario_data in structure.items():
            for discipline, disc_data in scenario_data['disciplines'].items():
                discipline_data.append({
                    'Scenario': scenario,
                    'Disiplin': discipline,
                    'Konstruksjon (A)': disc_data['total']['construction_a'],
                    'Drift (B)': disc_data['total']['operation_b'],
                    'Avslutning (C)': disc_data['total']['end_of_life_c'],
                    'Total GWP': disc_data['total']['total_gwp'],
                    'Antall rader': disc_data['total']['count']
                })

        if discipline_data:
            disc_df = pd.DataFrame(discipline_data)
            disc_df.to_excel(writer, sheet_name='👷 Disipliner', startrow=1, index=False)

            ws_disc = writer.sheets['👷 Disipliner']
            ws_disc['A1'] = 'DISIPLIN-ANALYSE'
            ws_disc['A1'].font = Font(size=14, bold=True, color='1565C0')

            # Header formatting
            for col in range(1, len(disc_df.columns) + 1):
                cell = ws_disc.cell(2, col)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='1565C0', end_color='1565C0', fill_type='solid')

        # ==============================================================================
        # SHEET 6: RAW DATA
        # ==============================================================================
        active_df = df[~df['excluded'] & df['mapped_scenario'].notna()].copy()
        export_cols = [
            'category', 'mapped_scenario', 'mapped_discipline', 'mapped_mmi_code',
            'construction_a', 'operation_b', 'end_of_life_c', 'total_gwp'
        ]
        active_export = active_df[export_cols].rename(columns={
            'category': 'Kategori',
            'mapped_scenario': 'Scenario',
            'mapped_discipline': 'Disiplin',
            'mapped_mmi_code': 'MMI',
            'construction_a': 'Konstruksjon (A)',
            'operation_b': 'Drift (B)',
            'end_of_life_c': 'Avslutning (C)',
            'total_gwp': 'Total GWP'
        })
        active_export.to_excel(writer, sheet_name='📋 Kartlagt data', index=False)

    excel_output.seek(0)
    return excel_output


# ==============================================================================
# PDF REPORT WITH VISUALIZATIONS
# ==============================================================================

class NumberedCanvas(canvas.Canvas):
    """Canvas with page numbers and header"""

    def __init__(self, *args, **kwargs):
        canvas.Canvas.__init__(self, *args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_number(num_pages)
            canvas.Canvas.showPage(self)
        canvas.Canvas.save(self)

    def draw_page_number(self, page_count):
        page_num = self._pageNumber
        text = f"Side {page_num} av {page_count}"
        self.setFont("Helvetica", 8)
        self.setFillColor(colors.grey)
        self.drawRightString(A4[0] - 2*cm, 1*cm, text)
        self.drawString(2*cm, 1*cm, f"Generert: {datetime.now().strftime('%d.%m.%Y')}")


def plotly_to_image(fig: go.Figure, width: int = 800, height: int = 500) -> io.BytesIO:
    """
    Convert plotly figure to image bytes.
    Uses kaleido if available, otherwise returns None.
    """
    try:
        img_bytes = fig.to_image(format="png", width=width, height=height, engine="kaleido")
        return io.BytesIO(img_bytes)
    except:
        # If kaleido not available, return None
        return None


def generate_pdf_report(df: pd.DataFrame, structure: Dict[str, Any],
                       scenario_summary: pd.DataFrame) -> io.BytesIO:
    """
    Generate enhanced PDF report with insights and visualizations.
    """
    pdf_output = io.BytesIO()
    doc = SimpleDocTemplate(pdf_output, pagesize=A4, topMargin=2*cm, bottomMargin=2.5*cm)

    elements = []
    styles = getSampleStyleSheet()

    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1565C0'),
        alignment=TA_CENTER,
        spaceAfter=10
    )

    subtitle_style = ParagraphStyle(
        'SubTitle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.grey,
        alignment=TA_CENTER,
        spaceAfter=30
    )

    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#1565C0'),
        spaceBefore=20,
        spaceAfter=10,
        borderPadding=(0, 0, 5, 0),
        borderColor=colors.HexColor('#1565C0'),
        borderWidth=2,
        borderRadius=0
    )

    body_style = ParagraphStyle(
        'Body',
        parent=styles['Normal'],
        fontSize=10,
        alignment=TA_JUSTIFY,
        spaceAfter=10
    )

    # ==============================================================================
    # TITLE PAGE
    # ==============================================================================
    elements.append(Spacer(1, 3*cm))
    elements.append(Paragraph("LCA SCENARIO MAPPING", title_style))
    elements.append(Paragraph("Klimagassanalyse og scenariosammenligning", subtitle_style))
    elements.append(Paragraph(
        f"Generert: {datetime.now().strftime('%d.%m.%Y kl. %H:%M')}",
        subtitle_style
    ))
    elements.append(Spacer(1, 2*cm))

    # Key metrics box
    insights = generate_insights(df, structure, scenario_summary)

    metrics_data = [
        ['NØKKELTALL', ''],
        ['Total rader', str(len(df))],
        ['Kartlagte rader', str(len(df[~df['excluded'] & df['mapped_scenario'].notna()]))],
        ['Ekskluderte rader', str(int(df['excluded'].sum()))],
        ['Antall scenarioer', str(len(structure))]
    ]

    metrics_table = Table(metrics_data, colWidths=[10*cm, 5*cm])
    metrics_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1565C0')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('BACKGROUND', (0, 1), (0, -1), colors.HexColor('#E3F2FD')),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('TOPPADDING', (0, 0), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey)
    ]))
    elements.append(metrics_table)

    elements.append(PageBreak())

    # ==============================================================================
    # EXECUTIVE SUMMARY
    # ==============================================================================
    elements.append(Paragraph("SAMMENDRAG", heading_style))

    for summary in insights['executive_summary']:
        elements.append(Paragraph(summary, body_style))

    elements.append(Spacer(1, 0.5*cm))

    # Key findings
    if insights['key_findings']:
        elements.append(Paragraph("HOVEDFUNN", heading_style))

        findings_data = [['Kategori', 'Verdi', 'Beskrivelse']]
        for finding in insights['key_findings']:
            findings_data.append([
                finding['title'],
                finding['value'],
                finding['description']
            ])

        findings_table = Table(findings_data, colWidths=[5*cm, 4*cm, 7*cm])
        findings_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1565C0')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
        ]))
        elements.append(findings_table)

    elements.append(Spacer(1, 0.5*cm))

    # Recommendations
    if insights['recommendations']:
        elements.append(Paragraph("ANBEFALINGER", heading_style))

        for i, rec in enumerate(insights['recommendations'], 1):
            elements.append(Paragraph(f"{i}. {rec}", body_style))

    elements.append(PageBreak())

    # ==============================================================================
    # SCENARIO SUMMARY
    # ==============================================================================
    elements.append(Paragraph("SCENARIO-OPPSUMMERING", heading_style))

    scenario_table_data = [['Scenario', 'Konstr. (A)', 'Drift (B)', 'Avsl. (C)', 'Total GWP']]
    for _, row in scenario_summary.iterrows():
        scenario_table_data.append([
            row['Scenario'],
            f"{row['Konstruksjon (A)']:,.0f}",
            f"{row['Drift (B)']:,.0f}",
            f"{row['Avslutning (C)']:,.0f}",
            f"{row['Total GWP']:,.0f}"
        ])

    scenario_table = Table(scenario_table_data, colWidths=[2*cm, 3.5*cm, 3.5*cm, 3.5*cm, 3.5*cm])
    scenario_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1565C0')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
    ]))
    elements.append(scenario_table)
    elements.append(Spacer(1, 1*cm))

    # Try to add chart
    try:
        from .visualizations import create_stacked_bar_chart

        chart_df = scenario_summary.copy()
        chart_df['construction_a'] = chart_df['Konstruksjon (A)']
        chart_df['operation_b'] = chart_df['Drift (B)']
        chart_df['end_of_life_c'] = chart_df['Avslutning (C)']

        fig = create_stacked_bar_chart(chart_df, 'Scenario', "GWP per Scenario")
        img_bytes = plotly_to_image(fig, width=600, height=400)

        if img_bytes:
            img = Image(img_bytes, width=15*cm, height=10*cm)
            elements.append(img)
    except Exception as e:
        elements.append(Paragraph(f"Kunne ikke generere diagram (installer 'kaleido' for diagrammer)", body_style))

    # ==============================================================================
    # C vs A COMPARISON
    # ==============================================================================
    if 'C' in structure and 'A' in structure:
        from .data_parser import compare_scenarios
        comparison = compare_scenarios(structure, 'A', 'C')

        if comparison:
            elements.append(PageBreak())
            elements.append(Paragraph("SCENARIO C vs A SAMMENLIGNING", heading_style))

            ratio = comparison['ratio']['total_gwp']
            diff = comparison['difference']['total_gwp']

            summary_text = f"Scenario C har {ratio:.1f}% av utslippene til Scenario A. "
            if ratio < 100:
                summary_text += f"Dette representerer en reduksjon på {abs(diff):,.0f} kg CO2e."
            else:
                summary_text += f"Dette representerer en økning på {abs(diff):,.0f} kg CO2e."

            elements.append(Paragraph(summary_text, body_style))
            elements.append(Spacer(1, 0.5*cm))

            comp_table_data = [['LCA-fase', 'Scenario A', 'Scenario C', 'Differanse', 'Ratio %']]
            for key, label in [
                ('total_gwp', 'Total GWP'),
                ('construction_a', 'Konstruksjon (A)'),
                ('operation_b', 'Drift (B)'),
                ('end_of_life_c', 'Avslutning (C)')
            ]:
                base_val = structure['A']['total'][key]
                comp_val = structure['C']['total'][key]
                phase_diff = comparison['difference'][key]
                phase_ratio = comparison['ratio'][key] if comparison['ratio'][key] else 0

                comp_table_data.append([
                    label,
                    f"{base_val:,.0f}",
                    f"{comp_val:,.0f}",
                    f"{phase_diff:+,.0f}",
                    f"{phase_ratio:.1f}%"
                ])

            comp_table = Table(comp_table_data, colWidths=[3.5*cm, 3*cm, 3*cm, 3*cm, 2.5*cm])
            comp_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1565C0')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
            ]))
            elements.append(comp_table)

    # ==============================================================================
    # MMI ANALYSIS
    # ==============================================================================
    if structure:
        elements.append(PageBreak())
        elements.append(Paragraph("MMI-ANALYSE", heading_style))

        from .visualizations import get_mmi_summary_stats

        for scenario in sorted(structure.keys()):
            mmi_stats = get_mmi_summary_stats(structure, scenario)

            if not mmi_stats.empty:
                elements.append(Paragraph(f"Scenario {scenario}", styles['Heading3']))

                mmi_table_data = [['MMI', 'Status', 'GWP (kg CO2e)', 'Andel (%)', 'Antall']]
                for _, row in mmi_stats.iterrows():
                    mmi_table_data.append([
                        row['MMI'],
                        row['Status'],
                        f"{row['GWP (kg CO2e)']:,.0f}",
                        f"{row['Andel (%)']:.1f}%",
                        str(row['Antall rader'])
                    ])

                mmi_table = Table(mmi_table_data, colWidths=[2*cm, 3*cm, 4*cm, 3*cm, 2*cm])
                mmi_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1565C0')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('ALIGN', (0, 0), (1, -1), 'LEFT'),
                    ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
                ]))
                elements.append(mmi_table)
                elements.append(Spacer(1, 0.5*cm))

    # Build PDF
    doc.build(elements, canvasmaker=NumberedCanvas)
    pdf_output.seek(0)
    return pdf_output
