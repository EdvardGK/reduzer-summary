# -*- coding: utf-8 -*-
"""
Enhanced Report Generation for LCA Scenario Mapping
Self-contained reports with charts, tables, and insights
PDF is master, Excel is support with matching structure
"""

import io
import logging
import pandas as pd
from datetime import datetime
from typing import Dict, Any, Optional, List, Tuple

# Setup logging
logger = logging.getLogger(__name__)

# PDF imports
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, Image, KeepTogether
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY
from reportlab.pdfgen import canvas
import plotly.graph_objects as go

# Excel formatting (no charts)
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows


# ==============================================================================
# INSIGHT GENERATION
# ==============================================================================

def generate_insights(df: pd.DataFrame, structure: Dict[str, Any],
                     scenario_summary: pd.DataFrame) -> Dict[str, Any]:
    """
    Generate key insights from the data.

    Returns:
        Dictionary with insights, findings, and recommendations
    """
    insights = {
        'executive_summary': [],
        'key_findings': [],
        'recommendations': [],
        'phase_analysis': []
    }

    # Mapping completeness
    total_rows = len(df)
    mapped_rows = len(df[~df['excluded'] & df['mapped_scenario'].notna()])
    excluded_rows = int(df['excluded'].sum())
    completeness = (mapped_rows / total_rows * 100) if total_rows > 0 else 0

    # Weighting stats
    weighted_rows = len(df[(~df['excluded']) & (df['weighting'] < 100)])
    avg_weighting = df[~df['excluded']]['weighting'].mean()

    insights['executive_summary'].append(
        f"Analysen dekker {total_rows} datapunkter, hvor {mapped_rows} ({completeness:.0f}%) "
        f"er kartlagt og {excluded_rows} er ekskludert. "
        f"Gjennomsnittlig vekting: {avg_weighting:.0f}% ({weighted_rows} rader har redusert vekt)."
    )

    # Scenario comparison
    if 'C' in structure and 'A' in structure:
        from .data_parser import compare_scenarios
        comparison = compare_scenarios(structure, 'A', 'C')

        if comparison:
            ratio = comparison['ratio']['total_gwp']
            diff = comparison['difference']['total_gwp']

            if ratio < 90:
                verdict = "betydelig reduksjon"
                color = "green"
            elif ratio < 100:
                verdict = "moderat reduksjon"
                color = "green"
            elif ratio < 110:
                verdict = "liten økning"
                color = "yellow"
            else:
                verdict = "betydelig økning"
                color = "red"

            insights['executive_summary'].append(
                f"Scenario C viser en {verdict} i klimagassutslipp sammenlignet med Scenario A "
                f"({ratio:.1f}%, {diff:+,.0f} kg CO2e)."
            )

            insights['key_findings'].append({
                'title': 'Scenario C vs A Sammenligning',
                'value': f"{ratio:.1f}%",
                'description': f"Differanse: {diff:+,.0f} kg CO2e",
                'color': color
            })

            # Phase analysis - detailed
            for key, label in [
                ('construction_a', 'Konstruksjon (A)'),
                ('operation_b', 'Drift (B)'),
                ('end_of_life_c', 'Avslutning (C)')
            ]:
                phase_ratio = comparison['ratio'][key]
                phase_diff = comparison['difference'][key]
                base_val = structure['A']['total'][key]
                comp_val = structure['C']['total'][key]

                insights['phase_analysis'].append({
                    'phase': label,
                    'scenario_a': base_val,
                    'scenario_c': comp_val,
                    'difference': phase_diff,
                    'ratio': phase_ratio
                })

                if phase_ratio and abs(100 - phase_ratio) > 10:
                    change = "reduksjon" if phase_ratio < 100 else "økning"
                    insights['key_findings'].append({
                        'title': f'{label}: {change.capitalize()}',
                        'value': f"{phase_ratio:.1f}%",
                        'description': f"{phase_diff:+,.0f} kg CO2e",
                        'color': 'green' if phase_ratio < 100 else 'red'
                    })

    # MMI distribution insights
    for scenario in structure.keys():
        from .visualizations import get_mmi_summary_stats
        mmi_stats = get_mmi_summary_stats(structure, scenario)

        if not mmi_stats.empty:
            # Find dominant MMI category
            dominant = mmi_stats.iloc[0]
            dominant_pct = dominant['Andel (%)']

            insights['key_findings'].append({
                'title': f'Scenario {scenario}: Dominerende MMI',
                'value': f"{dominant['MMI']} - {dominant['Status']}",
                'description': f"{dominant_pct:.1f}% av total GWP ({dominant['GWP (kg CO2e)']:,.0f} kg CO2e)",
                'color': 'blue'
            })

    # Discipline analysis
    if structure:
        for scenario, scenario_data in structure.items():
            disciplines = scenario_data['disciplines']
            if len(disciplines) > 0:
                # Find largest discipline contributor
                disc_totals = {disc: data['total']['total_gwp']
                              for disc, data in disciplines.items()}
                largest_disc = max(disc_totals, key=disc_totals.get)
                largest_gwp = disc_totals[largest_disc]
                scenario_total = scenario_data['total']['total_gwp']
                pct = (largest_gwp / scenario_total * 100) if scenario_total > 0 else 0

                insights['key_findings'].append({
                    'title': f'Scenario {scenario}: Største bidragsyter',
                    'value': largest_disc,
                    'description': f"{pct:.1f}% av total ({largest_gwp:,.0f} kg CO2e)",
                    'color': 'blue'
                })

    # Recommendations
    if 'C' in structure and 'A' in structure:
        comparison = compare_scenarios(structure, 'A', 'C')
        if comparison and comparison['ratio']['total_gwp'] > 110:
            insights['recommendations'].append(
                "Scenario C har betydelig høyere klimagassutslipp enn Scenario A. Vurder alternative løsninger "
                "for å redusere utslipp, spesielt i faser med størst økning."
            )
        elif comparison and comparison['ratio']['total_gwp'] < 90:
            insights['recommendations'].append(
                "Scenario C viser god reduksjon i klimagassutslipp. Dokumenter og implementer "
                "disse tiltakene for maksimal klimanytte."
            )
        elif comparison and 100 <= comparison['ratio']['total_gwp'] <= 110:
            insights['recommendations'].append(
                "Scenario C er tilnærmet likt Scenario A. Vurder ytterligere tiltak for å oppnå reduksjoner."
            )

    if completeness < 90:
        insights['recommendations'].append(
            f"Kartleggingsgrad er {completeness:.0f}%. Vurder å kartlegge flere datapunkter "
            "for mer komplett analyse."
        )

    return insights


# ==============================================================================
# EXCEL EXPORT - PURE DATA (NO CHARTS)
# ==============================================================================

def generate_excel_report(df: pd.DataFrame, structure: Dict[str, Any],
                         scenario_summary: pd.DataFrame) -> io.BytesIO:
    """
    Generate pure data Excel export - supplementary documentation.
    NO CHARTS - only clean, well-formatted data tables.

    Sheets:
    1. Metadata - Report info and summary statistics
    2. Scenario Summary - Aggregated by scenario
    3. Discipline Breakdown - All scenario x discipline combinations
    4. MMI Distribution - MMI categories per scenario
    5. C vs A Comparison - Detailed comparison if applicable
    6. Complete Dataset - All mapped rows with weighting
    """
    logger.info("Generating Excel data export...")
    excel_output = io.BytesIO()

    with pd.ExcelWriter(excel_output, engine='openpyxl') as writer:
        workbook = writer.book

        # ==============================================================================
        # SHEET 1: METADATA & SUMMARY STATISTICS
        # ==============================================================================
        ws_meta = workbook.create_sheet("📋 Metadata", 0)

        # Title
        ws_meta['A1'] = 'LCA SCENARIO MAPPING - DATA EXPORT'
        ws_meta['A1'].font = Font(size=16, bold=True, color='1565C0')
        ws_meta.merge_cells('A1:B1')

        ws_meta['A2'] = f"Generert: {datetime.now().strftime('%d.%m.%Y kl. %H:%M')}"
        ws_meta['A2'].font = Font(size=10, italic=True)
        ws_meta.merge_cells('A2:B2')

        # Summary stats
        row = 4
        ws_meta[f'A{row}'] = 'OPPSUMMERING'
        ws_meta[f'A{row}'].font = Font(size=12, bold=True, color='1565C0')
        row += 1

        # Weighting stats
        weighted_rows = len(df[(~df['excluded']) & (df['weighting'] < 100)])
        avg_weighting = df[~df['excluded']]['weighting'].mean()

        summary_data = [
            ['Total rader i datasett', len(df)],
            ['Kartlagte rader (aktive)', len(df[~df['excluded'] & df['mapped_scenario'].notna()])],
            ['Ekskluderte rader', int(df['excluded'].sum())],
            ['Antall scenarioer', len(structure)],
            ['Gjennomsnittlig vekting (%)', f"{avg_weighting:.1f}%"],
            ['Rader med redusert vekt', weighted_rows],
        ]

        for label, value in summary_data:
            ws_meta[f'A{row}'] = label
            ws_meta[f'B{row}'] = value
            ws_meta[f'A{row}'].font = Font(bold=True)
            ws_meta[f'A{row}'].fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
            row += 1

        row += 2

        # Column definitions
        ws_meta[f'A{row}'] = 'KOLONNEDEFINISJONER'
        ws_meta[f'A{row}'].font = Font(size=12, bold=True, color='1565C0')
        row += 1

        definitions = [
            ['Scenario', 'Prosjektalternativ (A, B, C, D)'],
            ['Disiplin', 'Fagområde (ARK, RIV, RIE, RIB, RIBp)'],
            ['MMI', 'Byggdel kategori (300=NY, 700=EKS, 800=GJEN, 900=RIVES)'],
            ['Vekting %', 'Prosentvis vekting av GWP (0-100%)'],
            ['Konstruksjon (A)', 'GWP i byggefase (kg CO2e)'],
            ['Drift (B)', 'GWP i driftsfase (kg CO2e)'],
            ['Avslutning (C)', 'GWP i avslutningsfase (kg CO2e)'],
            ['GWP basis', 'Total GWP uten vekting'],
            ['GWP vektet', 'Total GWP med vekting'],
        ]

        ws_meta[f'A{row}'] = 'Kolonne'
        ws_meta[f'B{row}'] = 'Beskrivelse'
        for col in ['A', 'B']:
            ws_meta[f'{col}{row}'].font = Font(bold=True)
            ws_meta[f'{col}{row}'].fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
        row += 1

        for col_name, desc in definitions:
            ws_meta[f'A{row}'] = col_name
            ws_meta[f'B{row}'] = desc
            row += 1

        # Column widths
        ws_meta.column_dimensions['A'].width = 25
        ws_meta.column_dimensions['B'].width = 60

        # ==============================================================================
        # SHEET 2: SCENARIO SUMMARY (NO CHARTS)
        # ==============================================================================
        scenario_summary.to_excel(writer, sheet_name='📊 Scenarioer', startrow=1, index=False)
        ws_scenarios = writer.sheets['📊 Scenarioer']

        # Header formatting
        for col in range(1, len(scenario_summary.columns) + 1):
            cell = ws_scenarios.cell(2, col)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='1565C0', end_color='1565C0', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')

        # Auto-size columns
        for col in ws_scenarios.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 50)
            ws_scenarios.column_dimensions[column].width = adjusted_width

        # ==============================================================================
        # SHEET 3: C vs A COMPARISON (NO CHARTS)
        # ==============================================================================
        if 'C' in structure and 'A' in structure:
            from .data_parser import compare_scenarios
            comparison = compare_scenarios(structure, 'A', 'C')

            if comparison:
                comp_data = []
                for key, label in [
                    ('total_gwp', 'Total GWP'),
                    ('construction_a', 'Konstruksjon (A)'),
                    ('operation_b', 'Drift (B)'),
                    ('end_of_life_c', 'Avslutning (C)')
                ]:
                    base_val = structure['A']['total'][key]
                    comp_val = structure['C']['total'][key]
                    diff = comparison['difference'][key]
                    ratio = comparison['ratio'][key] if comparison['ratio'][key] else 0

                    # Verdict
                    if ratio < 90:
                        verdict = "Stor reduksjon ✓"
                    elif ratio < 100:
                        verdict = "Reduksjon ✓"
                    elif ratio < 110:
                        verdict = "Liten økning ⚠"
                    else:
                        verdict = "Stor økning ✗"

                    comp_data.append({
                        'LCA-fase': label,
                        'Scenario A (kg CO2e)': base_val,
                        'Scenario C (kg CO2e)': comp_val,
                        'Differanse (C-A)': diff,
                        'Ratio (C/A) %': ratio,
                        'Vurdering': verdict
                    })

                comp_df = pd.DataFrame(comp_data)
                comp_df.to_excel(writer, sheet_name='🔄 C vs A', startrow=1, index=False)
                ws_comp = writer.sheets['🔄 C vs A']

                # Header formatting
                for col in range(1, len(comp_df.columns) + 1):
                    cell = ws_comp.cell(2, col)
                    cell.font = Font(bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color='1565C0', end_color='1565C0', fill_type='solid')
                    cell.alignment = Alignment(horizontal='center')

                # Conditional formatting on ratio column (column 5)
                for row_idx in range(3, len(comp_data) + 3):
                    ratio_cell = ws_comp.cell(row_idx, 5)
                    ratio_val = ratio_cell.value

                    if ratio_val and ratio_val < 100:
                        ratio_cell.fill = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')
                    elif ratio_val:
                        ratio_cell.fill = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')

                # Auto-size columns
                for col in ws_comp.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    adjusted_width = min(max_length + 2, 50)
                    ws_comp.column_dimensions[column].width = adjusted_width

        # ==============================================================================
        # SHEET 4: MMI DISTRIBUTION (NO CHARTS)
        # ==============================================================================
        from .visualizations import get_mmi_summary_stats
        mmi_dist_data = []
        for scenario in sorted(structure.keys()):
            mmi_stats = get_mmi_summary_stats(structure, scenario)
            if not mmi_stats.empty:
                mmi_stats['Scenario'] = scenario
                mmi_dist_data.append(mmi_stats)

        if mmi_dist_data:
            mmi_dist_df = pd.concat(mmi_dist_data, ignore_index=True)
            mmi_dist_df = mmi_dist_df[['Scenario', 'MMI', 'Status', 'GWP (kg CO2e)', 'Andel (%)', 'Antall rader']]
            mmi_dist_df.to_excel(writer, sheet_name='🏗️ MMI', startrow=1, index=False)

            ws_mmi = writer.sheets['🏗️ MMI']

            # Header formatting
            for col in range(1, len(mmi_dist_df.columns) + 1):
                cell = ws_mmi.cell(2, col)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='1565C0', end_color='1565C0', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')

            # Auto-size columns
            for col in ws_mmi.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = min(max_length + 2, 50)
                ws_mmi.column_dimensions[column].width = adjusted_width

        # ==============================================================================
        # SHEET 5: DISCIPLINE BREAKDOWN (NO CHARTS)
        # ==============================================================================
        discipline_data = []
        for scenario, scenario_data in structure.items():
            for discipline, disc_data in scenario_data['disciplines'].items():
                discipline_data.append({
                    'Scenario': scenario,
                    'Disiplin': discipline,
                    'Konstruksjon (A) (kg CO2e)': disc_data['total']['construction_a'],
                    'Drift (B) (kg CO2e)': disc_data['total']['operation_b'],
                    'Avslutning (C) (kg CO2e)': disc_data['total']['end_of_life_c'],
                    'Total GWP (kg CO2e)': disc_data['total']['total_gwp'],
                    'Antall rader': disc_data['total']['count']
                })

        if discipline_data:
            # Sort by total GWP descending
            disc_df = pd.DataFrame(discipline_data)
            disc_df = disc_df.sort_values('Total GWP (kg CO2e)', ascending=False)
            disc_df.to_excel(writer, sheet_name='👷 Disipliner', startrow=1, index=False)

            ws_disc = writer.sheets['👷 Disipliner']

            # Header formatting
            for col in range(1, len(disc_df.columns) + 1):
                cell = ws_disc.cell(2, col)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='1565C0', end_color='1565C0', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')

            # Auto-size columns
            for col in ws_disc.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = min(max_length + 2, 50)
                ws_disc.column_dimensions[column].width = adjusted_width

        # ==============================================================================
        # SHEET 6: COMPLETE DATASET WITH ALL FIELDS
        # ==============================================================================
        active_df = df[~df['excluded'] & df['mapped_scenario'].notna()].copy()

        # Select and rename columns for export
        export_cols = [
            'category',
            'mapped_scenario',
            'mapped_discipline',
            'mapped_mmi_code',
            'weighting',
            'construction_a',
            'operation_b',
            'end_of_life_c',
            'total_gwp_base',
            'total_gwp'
        ]

        active_export = active_df[export_cols].rename(columns={
            'category': 'Kategori',
            'mapped_scenario': 'Scenario',
            'mapped_discipline': 'Disiplin',
            'mapped_mmi_code': 'MMI',
            'weighting': 'Vekting (%)',
            'construction_a': 'Konstruksjon (A) (kg CO2e)',
            'operation_b': 'Drift (B) (kg CO2e)',
            'end_of_life_c': 'Avslutning (C) (kg CO2e)',
            'total_gwp_base': 'GWP basis (kg CO2e)',
            'total_gwp': 'GWP vektet (kg CO2e)'
        })

        # Sort by scenario, then discipline, then MMI
        active_export = active_export.sort_values(['Scenario', 'Disiplin', 'MMI'])
        active_export.to_excel(writer, sheet_name='📋 Komplett datasett', startrow=1, index=False)

        ws_data = writer.sheets['📋 Komplett datasett']

        # Header formatting
        for col in range(1, len(active_export.columns) + 1):
            cell = ws_data.cell(2, col)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='1565C0', end_color='1565C0', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')

        # Auto-size columns
        for col in ws_data.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 50)
            ws_data.column_dimensions[column].width = adjusted_width

    excel_output.seek(0)
    logger.info("Excel export completed successfully")
    return excel_output


# ==============================================================================
# PDF REPORT - MASTER DOCUMENT
# ==============================================================================

class NumberedCanvas(canvas.Canvas):
    """Canvas with page numbers and footer"""

    def __init__(self, *args, **kwargs):
        canvas.Canvas.__init__(self, *args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_number(num_pages)
            canvas.Canvas.showPage(self)
        canvas.Canvas.save(self)

    def draw_page_number(self, page_count):
        page_num = self._pageNumber
        text = f"Side {page_num} av {page_count}"
        self.setFont("Helvetica", 8)
        self.setFillColor(colors.grey)
        self.drawRightString(A4[0] - 2*cm, 1*cm, text)
        self.drawString(2*cm, 1*cm, f"Generert: {datetime.now().strftime('%d.%m.%Y')}")


def plotly_to_image(fig: go.Figure, width: int = 800, height: int = 500) -> Optional[io.BytesIO]:
    """
    Convert plotly figure to image bytes.
    Uses kaleido if available, otherwise returns None.
    """
    try:
        logger.info(f"Converting plotly figure to image: {width}x{height}px")
        img_bytes = fig.to_image(format="png", width=width, height=height, scale=2, engine="kaleido")
        logger.info("Successfully converted plotly figure to image")
        return io.BytesIO(img_bytes)
    except ImportError as e:
        logger.error(f"Kaleido not installed or import failed: {e}")
        return None
    except ValueError as e:
        logger.error(f"Invalid figure or parameters: {e}")
        return None
    except TimeoutError as e:
        logger.warning(f"Kaleido timeout (first call can take 5+ minutes): {e}")
        return None
    except Exception as e:
        logger.error(f"Unexpected error converting plotly to image: {type(e).__name__}: {e}", exc_info=True)
        return None


def create_metric_box(label: str, value: str, color: str = 'blue', width: float = 6) -> Table:
    """
    Create a visual callout box for key metrics.

    Args:
        label: Metric label (e.g., "C vs A Ratio")
        value: Metric value (e.g., "95.2%")
        color: Color theme - 'green', 'red', 'blue', 'yellow'
        width: Width in cm

    Returns:
        ReportLab Table object
    """
    color_map = {
        'green': colors.HexColor('#4CAF50'),
        'red': colors.HexColor('#F44336'),
        'blue': colors.HexColor('#1565C0'),
        'yellow': colors.HexColor('#FFC107'),
        'grey': colors.HexColor('#757575')
    }

    bg_color = color_map.get(color, colors.HexColor('#1565C0'))

    data = [[label], [value]]
    table = Table(data, colWidths=[width*cm], rowHeights=[0.8*cm, 1.8*cm])
    table.setStyle(TableStyle([
        # Header row
        ('BACKGROUND', (0, 0), (-1, 0), bg_color),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),

        # Value row
        ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#F5F5F5')),
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 1), (-1, 1), 24),
        ('ALIGN', (0, 1), (-1, 1), 'CENTER'),
        ('VALIGN', (0, 1), (-1, 1), 'MIDDLE'),

        # Border
        ('BOX', (0, 0), (-1, -1), 2, bg_color),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    return table


def create_section_divider() -> Table:
    """
    Create a visual divider line between sections.

    Returns:
        ReportLab Table object
    """
    divider = Table([['']], colWidths=[16*cm], rowHeights=[0.2*cm])
    divider.setStyle(TableStyle([
        ('LINEABOVE', (0, 0), (-1, 0), 3, colors.HexColor('#1565C0')),
        ('LINEBELOW', (0, 0), (-1, 0), 1, colors.HexColor('#E3F2FD'))
    ]))
    return divider


def test_kaleido() -> bool:
    """
    Test if Kaleido is working correctly.
    Returns True if successful, False otherwise.
    """
    try:
        logger.info("Testing Kaleido installation...")
        test_fig = go.Figure(go.Bar(x=['Test'], y=[1]))
        test_bytes = test_fig.to_image(format="png", engine="kaleido", width=100, height=100)
        logger.info("✓ Kaleido is working correctly")
        return True
    except ImportError as e:
        logger.error(f"✗ Kaleido import failed: {e}")
        return False
    except Exception as e:
        logger.error(f"✗ Kaleido test failed: {type(e).__name__}: {e}")
        return False


def generate_pdf_report(df: pd.DataFrame, structure: Dict[str, Any],
                       scenario_summary: pd.DataFrame) -> io.BytesIO:
    """
    Generate comprehensive PDF report (MASTER DOCUMENT).
    Self-contained with all charts, tables, and insights.
    """
    logger.info("Starting PDF report generation...")

    # Test Kaleido before generating report
    kaleido_available = test_kaleido()
    if not kaleido_available:
        logger.warning("Charts will not be included in PDF - Kaleido is not available")

    pdf_output = io.BytesIO()
    doc = SimpleDocTemplate(pdf_output, pagesize=A4, topMargin=2*cm, bottomMargin=2.5*cm)

    elements = []
    styles = getSampleStyleSheet()

    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1565C0'),
        alignment=TA_CENTER,
        spaceAfter=10
    )

    subtitle_style = ParagraphStyle(
        'SubTitle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.grey,
        alignment=TA_CENTER,
        spaceAfter=20
    )

    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#1565C0'),
        spaceBefore=15,
        spaceAfter=8,
        borderPadding=(0, 0, 5, 0),
        borderColor=colors.HexColor('#1565C0'),
        borderWidth=2,
        borderRadius=0
    )

    heading3_style = ParagraphStyle(
        'Heading3Custom',
        parent=styles['Heading3'],
        fontSize=11,
        textColor=colors.HexColor('#1565C0'),
        spaceBefore=10,
        spaceAfter=6
    )

    body_style = ParagraphStyle(
        'Body',
        parent=styles['Normal'],
        fontSize=9,
        alignment=TA_JUSTIFY,
        spaceAfter=6,
        leading=12,
        wordWrap='CJK'
    )

    # Generate insights
    insights = generate_insights(df, structure, scenario_summary)

    # ==============================================================================
    # TITLE PAGE
    # ==============================================================================
    elements.append(Spacer(1, 2*cm))
    elements.append(Paragraph("LCA SCENARIO MAPPING", title_style))
    elements.append(Paragraph("Klimagassanalyse og scenariosammenligning", subtitle_style))
    elements.append(Paragraph(
        f"Generert: {datetime.now().strftime('%d.%m.%Y kl. %H:%M')}",
        subtitle_style
    ))
    elements.append(Spacer(1, 1*cm))

    # Key metrics box
    weighted_rows = len(df[(~df['excluded']) & (df['weighting'] < 100)])
    avg_weighting = df[~df['excluded']]['weighting'].mean()

    metrics_data = [
        ['NØKKELTALL', ''],
        ['Total rader', str(len(df))],
        ['Kartlagte rader', str(len(df[~df['excluded'] & df['mapped_scenario'].notna()]))],
        ['Ekskluderte rader', str(int(df['excluded'].sum()))],
        ['Antall scenarioer', str(len(structure))],
        ['Gj.snitt vekting', f"{avg_weighting:.0f}%"],
        ['Rader m/redusert vekt', str(weighted_rows)]
    ]

    metrics_table = Table(metrics_data, colWidths=[9*cm, 4*cm])
    metrics_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1565C0')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('BACKGROUND', (0, 1), (0, -1), colors.HexColor('#E3F2FD')),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')])
    ]))
    elements.append(metrics_table)
    elements.append(Spacer(1, 0.8*cm))

    # ==============================================================================
    # VISUAL KPIs - KEY METRICS BOXES
    # ==============================================================================
    if 'C' in structure and 'A' in structure:
        from .data_parser import compare_scenarios
        comparison = compare_scenarios(structure, 'A', 'C')

        if comparison and comparison['ratio']['total_gwp']:
            ratio = comparison['ratio']['total_gwp']
            diff = comparison['difference']['total_gwp']

            # Determine color based on ratio
            if ratio < 95:
                color = 'green'
            elif ratio < 105:
                color = 'yellow'
            else:
                color = 'red'

            # Create 3 metric boxes side by side
            box1 = create_metric_box("C vs A Ratio", f"{ratio:.1f}%", color, width=5)
            box2 = create_metric_box("Differanse", f"{diff:+,.0f} kg", color, width=5)
            box3 = create_metric_box("Datakvalitet", f"{completeness:.0f}%", 'blue', width=5)

            boxes_table = Table([[box1, box2, box3]], colWidths=[5.2*cm, 5.2*cm, 5.2*cm])
            boxes_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            elements.append(boxes_table)
            elements.append(Spacer(1, 0.6*cm))

            # Add verdict paragraph
            if ratio < 100:
                verdict = f"✓ Scenario C har {abs(100-ratio):.1f}% lavere klimagassutslipp enn Scenario A"
                verdict_color = colors.HexColor('#4CAF50')
            else:
                verdict = f"✗ Scenario C har {ratio-100:.1f}% høyere klimagassutslipp enn Scenario A"
                verdict_color = colors.HexColor('#F44336')

            verdict_style = ParagraphStyle(
                'Verdict',
                parent=body_style,
                fontSize=12,
                textColor=verdict_color,
                alignment=TA_CENTER,
                spaceAfter=12,
                fontName='Helvetica-Bold'
            )
            elements.append(Paragraph(verdict, verdict_style))

    elements.append(create_section_divider())
    elements.append(Spacer(1, 0.4*cm))

    # ==============================================================================
    # EXECUTIVE SUMMARY - CONDENSED
    # ==============================================================================
    elements.append(Paragraph("SAMMENDRAG", heading_style))

    # Convert summary to bullet points for scannability
    for summary in insights['executive_summary']:
        bullet_text = f"• {summary}"
        elements.append(Paragraph(bullet_text, body_style))

    elements.append(Spacer(1, 0.3*cm))

    # Key findings table
    if insights['key_findings']:
        findings_elements = []
        findings_elements.append(Paragraph("HOVEDFUNN", heading_style))

        findings_data = [['Kategori', 'Verdi', 'Beskrivelse']]
        for finding in insights['key_findings']:
            findings_data.append([
                finding['title'],
                finding['value'],
                finding['description']
            ])

        findings_table = Table(findings_data, colWidths=[4.5*cm, 3.5*cm, 7*cm])
        findings_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1565C0')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
            ('WORDWRAP', (0, 0), (-1, -1), True)
        ]))
        findings_elements.append(findings_table)

        # Keep findings together on same page
        elements.append(KeepTogether(findings_elements))

    elements.append(Spacer(1, 0.3*cm))

    # Recommendations
    if insights['recommendations']:
        elements.append(Paragraph("ANBEFALINGER", heading3_style))
        for i, rec in enumerate(insights['recommendations'], 1):
            elements.append(Paragraph(f"• {rec}", body_style))
        elements.append(Spacer(1, 0.2*cm))

    # Add hero chart on page 1 if comparison available
    if 'C' in structure and 'A' in structure:
        from .data_parser import compare_scenarios
        comparison = compare_scenarios(structure, 'A', 'C')

        if comparison and kaleido_available:
            try:
                elements.append(create_section_divider())
                elements.append(Spacer(1, 0.3*cm))
                elements.append(Paragraph("SCENARIO C vs A - VISUELL SAMMENLIGNING", heading_style))

                from .visualizations import create_comparison_chart
                fig_diff = create_comparison_chart(comparison, 'difference')
                img_diff = plotly_to_image(fig_diff, width=650, height=320)

                if img_diff:
                    img = Image(img_diff, width=15*cm, height=7.5*cm)
                    elements.append(img)
                else:
                    logger.warning("Failed to generate comparison chart for page 1")
            except Exception as e:
                logger.error(f"Error adding hero chart to page 1: {e}")

    elements.append(PageBreak())

    # ==============================================================================
    # SCENARIO SUMMARY WITH CHART
    # ==============================================================================
    elements.append(create_section_divider())
    elements.append(Spacer(1, 0.4*cm))

    scenario_elements = []
    scenario_elements.append(Paragraph("SCENARIO-OPPSUMMERING", heading_style))

    scenario_table_data = [['Scenario', 'Konstr. (A)', 'Drift (B)', 'Avsl. (C)', 'Total GWP']]
    for _, row in scenario_summary.iterrows():
        scenario_table_data.append([
            row['Scenario'],
            f"{row['Konstruksjon (A)']:,.0f}",
            f"{row['Drift (B)']:,.0f}",
            f"{row['Avslutning (C)']:,.0f}",
            f"{row['Total GWP']:,.0f}"
        ])

    scenario_table = Table(scenario_table_data, colWidths=[1.5*cm, 3*cm, 3*cm, 3*cm, 3.5*cm])
    scenario_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1565C0')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (0, -1), 'CENTER'),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')])
    ]))
    scenario_elements.append(scenario_table)
    scenario_elements.append(Spacer(1, 0.3*cm))

    # Add chart
    try:
        from .visualizations import create_stacked_bar_chart

        chart_df = scenario_summary.copy()
        chart_df['construction_a'] = chart_df['Konstruksjon (A)']
        chart_df['operation_b'] = chart_df['Drift (B)']
        chart_df['end_of_life_c'] = chart_df['Avslutning (C)']

        fig = create_stacked_bar_chart(chart_df, 'Scenario', "GWP per Scenario")
        img_bytes = plotly_to_image(fig, width=650, height=380)

        if img_bytes:
            img = Image(img_bytes, width=14*cm, height=8*cm)
            scenario_elements.append(img)
        else:
            logger.warning("Scenario stacked bar chart not available")
            if not kaleido_available:
                scenario_elements.append(Paragraph(
                    "⚠ Diagram krever Kaleido. Installer med: pip install kaleido",
                    body_style
                ))
    except Exception as e:
        logger.error(f"Error generating scenario chart: {e}")
        scenario_elements.append(Paragraph("⚠ Feil ved generering av diagram", body_style))

    elements.extend(scenario_elements)
    elements.append(PageBreak())

    # ==============================================================================
    # C vs A COMPARISON WITH CHARTS
    # ==============================================================================
    elements.append(create_section_divider())
    elements.append(Spacer(1, 0.4*cm))

    if 'C' in structure and 'A' in structure:
        from .data_parser import compare_scenarios
        comparison = compare_scenarios(structure, 'A', 'C')

        if comparison:
            comp_elements = []
            comp_elements.append(Paragraph("SCENARIO C vs A SAMMENLIGNING", heading_style))

            ratio = comparison['ratio']['total_gwp']
            diff = comparison['difference']['total_gwp']

            summary_text = f"Scenario C har {ratio:.1f}% av utslippene til Scenario A. "
            if ratio < 100:
                summary_text += f"Dette representerer en reduksjon på {abs(diff):,.0f} kg CO2e ({abs(100-ratio):.1f}% lavere)."
            else:
                summary_text += f"Dette representerer en økning på {abs(diff):,.0f} kg CO2e ({ratio-100:.1f}% høyere)."

            comp_elements.append(Paragraph(summary_text, body_style))
            comp_elements.append(Spacer(1, 0.3*cm))

            # Comparison table
            comp_table_data = [['LCA-fase', 'Scenario A', 'Scenario C', 'Differanse', 'Ratio %', 'Vurdering']]
            for key, label in [
                ('total_gwp', 'Total GWP'),
                ('construction_a', 'Konstruksjon (A)'),
                ('operation_b', 'Drift (B)'),
                ('end_of_life_c', 'Avslutning (C)')
            ]:
                base_val = structure['A']['total'][key]
                comp_val = structure['C']['total'][key]
                phase_diff = comparison['difference'][key]
                phase_ratio = comparison['ratio'][key] if comparison['ratio'][key] else 0

                if phase_ratio < 90:
                    verdict = "Stor reduksjon ✓"
                elif phase_ratio < 100:
                    verdict = "Reduksjon ✓"
                elif phase_ratio < 110:
                    verdict = "Liten økning ⚠"
                else:
                    verdict = "Stor økning ✗"

                comp_table_data.append([
                    label,
                    f"{base_val:,.0f}",
                    f"{comp_val:,.0f}",
                    f"{phase_diff:+,.0f}",
                    f"{phase_ratio:.1f}%",
                    verdict
                ])

            comp_table = Table(comp_table_data, colWidths=[3*cm, 2.5*cm, 2.5*cm, 2.5*cm, 2*cm, 3.5*cm])
            comp_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1565C0')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
            ]))
            comp_elements.append(comp_table)
            comp_elements.append(Spacer(1, 0.3*cm))

            # Add comparison charts side by side if possible
            try:
                from .visualizations import create_comparison_chart

                # Difference chart
                fig_diff = create_comparison_chart(comparison, 'difference')
                img_diff = plotly_to_image(fig_diff, width=700, height=400)

                if img_diff:
                    comp_elements.append(Paragraph("Differanse (kg CO2e)", heading3_style))
                    comp_elements.append(Image(img_diff, width=14*cm, height=8*cm))
                    comp_elements.append(Spacer(1, 0.3*cm))

                # Ratio chart
                fig_ratio = create_comparison_chart(comparison, 'ratio')
                img_ratio = plotly_to_image(fig_ratio, width=700, height=400)

                if img_ratio:
                    comp_elements.append(Paragraph("Ratio (%)", heading3_style))
                    comp_elements.append(Image(img_ratio, width=14*cm, height=8*cm))

            except Exception as e:
                pass

            elements.extend(comp_elements)
            elements.append(PageBreak())

    # ==============================================================================
    # MMI ANALYSIS WITH CHARTS
    # ==============================================================================
    elements.append(create_section_divider())
    elements.append(Spacer(1, 0.4*cm))

    if structure:
        elements.append(Paragraph("MMI-ANALYSE", heading_style))

        from .visualizations import get_mmi_summary_stats, create_mmi_distribution_pie, create_mmi_distribution_by_discipline

        for scenario in sorted(structure.keys()):
            mmi_stats = get_mmi_summary_stats(structure, scenario)

            if not mmi_stats.empty:
                mmi_elements = []
                mmi_elements.append(Paragraph(f"Scenario {scenario}", heading3_style))

                # MMI table
                mmi_table_data = [['MMI', 'Status', 'GWP (kg CO2e)', 'Andel (%)', 'Antall']]
                for _, row in mmi_stats.iterrows():
                    mmi_table_data.append([
                        row['MMI'],
                        row['Status'],
                        f"{row['GWP (kg CO2e)']:,.0f}",
                        f"{row['Andel (%)']:.1f}%",
                        str(row['Antall rader'])
                    ])

                mmi_table = Table(mmi_table_data, colWidths=[2*cm, 3*cm, 3.5*cm, 2.5*cm, 2*cm])
                mmi_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1565C0')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('ALIGN', (0, 0), (1, -1), 'LEFT'),
                    ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
                    ('TOPPADDING', (0, 0), (-1, -1), 5),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
                ]))
                mmi_elements.append(mmi_table)
                mmi_elements.append(Spacer(1, 0.3*cm))

                # Add MMI pie chart
                try:
                    fig_pie = create_mmi_distribution_pie(structure, scenario)
                    img_pie = plotly_to_image(fig_pie, width=600, height=400)

                    if img_pie:
                        mmi_elements.append(Image(img_pie, width=12*cm, height=8*cm))
                        mmi_elements.append(Spacer(1, 0.3*cm))
                except Exception as e:
                    pass

                # Add MMI by discipline chart
                try:
                    fig_disc = create_mmi_distribution_by_discipline(structure, scenario)
                    img_disc = plotly_to_image(fig_disc, width=700, height=400)

                    if img_disc:
                        mmi_elements.append(Paragraph("MMI-fordeling per disiplin", heading3_style))
                        mmi_elements.append(Image(img_disc, width=14*cm, height=8*cm))
                except Exception as e:
                    pass

                elements.extend(mmi_elements)
                elements.append(Spacer(1, 0.5*cm))

    # ==============================================================================
    # DISCIPLINE BREAKDOWN
    # ==============================================================================
    elements.append(PageBreak())
    elements.append(create_section_divider())
    elements.append(Spacer(1, 0.4*cm))
    elements.append(Paragraph("DISIPLIN-ANALYSE", heading_style))

    discipline_data = []
    for scenario, scenario_data in structure.items():
        for discipline, disc_data in scenario_data['disciplines'].items():
            discipline_data.append({
                'Scenario': scenario,
                'Disiplin': discipline,
                'Konstruksjon (A)': disc_data['total']['construction_a'],
                'Drift (B)': disc_data['total']['operation_b'],
                'Avslutning (C)': disc_data['total']['end_of_life_c'],
                'Total GWP': disc_data['total']['total_gwp'],
                'Antall': disc_data['total']['count']
            })

    if discipline_data:
        disc_df = pd.DataFrame(discipline_data)

        # Discipline table
        disc_table_data = [['Scenario', 'Disiplin', 'Konstr.', 'Drift', 'Avsl.', 'Total', 'Ant.']]
        for _, row in disc_df.iterrows():
            disc_table_data.append([
                row['Scenario'],
                row['Disiplin'],
                f"{row['Konstruksjon (A)']:,.0f}",
                f"{row['Drift (B)']:,.0f}",
                f"{row['Avslutning (C)']:,.0f}",
                f"{row['Total GWP']:,.0f}",
                str(row['Antall'])
            ])

        disc_table = Table(disc_table_data, colWidths=[1.5*cm, 2.5*cm, 2.5*cm, 2.5*cm, 2.5*cm, 2.5*cm, 1.5*cm])
        disc_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1565C0')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (0, 0), (1, -1), 'LEFT'),
            ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
        ]))
        elements.append(disc_table)
        elements.append(Spacer(1, 0.3*cm))

        # Add discipline chart per scenario
        try:
            from .visualizations import create_discipline_comparison_chart

            for scenario in sorted(structure.keys()):
                fig_disc_scenario = create_discipline_comparison_chart(structure, scenario)
                img_disc_scenario = plotly_to_image(fig_disc_scenario, width=700, height=450)

                if img_disc_scenario:
                    elements.append(Paragraph(f"Disipliner - Scenario {scenario}", heading3_style))
                    elements.append(Image(img_disc_scenario, width=14*cm, height=9*cm))
                    elements.append(Spacer(1, 0.3*cm))
        except Exception as e:
            pass

    # Build PDF
    doc.build(elements, canvasmaker=NumberedCanvas)
    pdf_output.seek(0)
    return pdf_output
